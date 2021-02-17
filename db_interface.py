#! D:\APP\python\python.exe
# -*- coding:utf-8 -*-

import pymysql
import openpyxl  # 可以操作xlsx文件(2007以上)
from db_info import *


def read_excel(filename, table_index):
    data = []
    row_data = []
    # 1. 获取工作簿对象
    workbook = openpyxl.load_workbook(filename)
    # 2. 获取所有sheet的名称
    sheet_names = workbook.sheetnames
    # print(sheet_names)
    # 3. 获取一个特定sheet
    # worksheet = workbook["表1"]
    # worksheet = workbook[sheet_names[0]]             # 通过表名
    worksheet = workbook.worksheets[table_index]  # 通过索引
    # 4. 获取工作表的属性
    sheet_name = worksheet.title
    rows = worksheet.max_row
    columns = worksheet.max_column
    # print(sheet_name, rows, columns)
    # 5. 获取数据(sheet.rows 行生成器;sheet.columns列生成器)
    for row in worksheet.rows:
        for cell in row:
            row_data.append(cell.value)
        data.append(row_data)
        row_data = []

    # for col in worksheet.columns:
    #     for cell in col:
    #         print(cell.value, end=" ")
    #     print()
    # 6.获取特定行或特定列的方式
    # " 方法1: 索引从0开始"
    # for row in list(worksheet.rows)[0:3]:
    #     for cell in row[0:3]:
    #         print(cell.value, end=" ")
    #     print()
    # " 方法2: 索引从1开始"
    # for i in range(1, 4):
    #     for j in range(1, 4):
    #         print(worksheet.cell(row=i, column=j).value, end=" ")
    #     print()
    # 7. 获取某个单元格的数据
    # print(worksheet['A2'].value)
    return data


def write_excel(filename, sheet_index, data):
    sheet_name_dict ={1:"表1(图书信息表)", 2:"表2(顾客信息表)", 3:"表3(订单信息表)"}
    workbook = openpyxl.Workbook()       # 创建工作簿
    export_num_flag = 0
    for i in range(len(sheet_name_dict)):  # i是0,1... 表示第几个表
        sheet_name = sheet_name_dict[i+1]    # 肯定要创建3个表
        if i == 0:
            worksheet = workbook.active
            worksheet.title = sheet_name
        else:
            worksheet = workbook.create_sheet(title=sheet_name)
        try:
            sheet_index_real = sheet_index[export_num_flag]
            tmp = data[sheet_index_real]
            if i < sheet_index_real:
                pass
            elif i == sheet_index_real:
                row = len(tmp)
                column = len(tmp[0])
                export_num_flag += 1
                for i in range(row):
                    for j in range(column):
                        worksheet.cell(i + 1, j + 1).value = tmp[i][j]  # sheet索引从1开始算起
        except:
            pass
    # 4. 保存
    workbook.save(filename)
    return True


class Mysql_db():
    def __init__(self, **db_info):
        """
        数据库连接字典
        :param db_info: host, port, db, user, passwd...
        """
        self._db_info = db_info
        self.table_dict = {0: "book_item", 1: "customer_info", 2: "order_info"}

    def __enter__(self):
        """数据库连接"""
        self.conn = pymysql.connect(**self._db_info)
        self.cur = self.conn.cursor()
        return self

    def __exit__(self, exc_type, exc_value, exc_tb):
        """
        游标和数据库优雅关闭
        :param exc_type: 异常类型
        :param exc_value: 异常值
        :param exc_tb: 异常的错误栈信息
        :return:
        """
        try:
            self.conn.commit()  # 提交
            print("commit successfully!")
        except:
            self.conn.rollback()  # 如果发生错误时回滚
            print("your commit fails")
        finally:
            self.cur.close()
            self.conn.close()


    def insert_bookitem(self, ISBN, bname, price, inventory=1):
        sql = "INSERT INTO book_item(ISBN, bname, \
                price, inventory) \
                VALUES('%s', '%s', '%s', '%s')" % \
              (ISBN, bname, price, inventory)
        self.cur.execute(sql)

    def insert_customerinfo(self, uid, uname, card_class, rental_time, expire_date):
        sql = "INSERT INTO customer_info(uid, uname, \
                card_class, rental_time, expire_date) \
                VALUES('%s', '%s', '%s', '%s', '%s')" % \
              (uid, uname, card_class, rental_time, expire_date)
        self.cur.execute(sql)

    def insert_orderinfo_start(self, uid, uname, bname, ISBN, start_time, end_time=None):
        if end_time == None or end_time == "":
            sql = "INSERT INTO order_info(uid, uname, ISBN, bname, start_time\
                  ) VALUES('%s', '%s', '%s', '%s', '%s')" % \
                  (uid, uname, bname, ISBN, start_time)
        else:
            sql = "INSERT INTO order_info(uid, uname, ISBN, bname, start_time, end_time\
                  ) VALUES('%s', '%s', '%s', '%s', '%s', '%s')" % \
                  (uid, uname, bname, ISBN, start_time, end_time)
        self.cur.execute(sql)

    def commit_db(self, data, table_index):
        """
       先把表的数据删除, 然后再根据data插入数据(简单实现)
       :param data: [(xxx, xxx, xxx, xxx),(xxx,xxx,xxx,xxx)...]
       :param table_index: 1,2,3
       :return: None
       """
        table_name = self.table_dict[table_index]  # 先获得表名称
        delete_sql = "Truncate Table %s" % (table_name)  # 删除表数据,但是保留表的结构;类似delete不加from;drop全部删除
        self.cur.execute(delete_sql)
        if table_index == 0:
            for row in data:
                self.insert_bookitem(row[0], row[1], row[2], row[3])
        elif table_index == 1:
            for row in data:
                self.insert_customerinfo(row[0], row[1], row[2], get_date_string(row[3]), row[4])
        elif table_index == 2:
            for row in data:
                self.insert_orderinfo_start(row[0], row[1], row[2], row[3], get_date_string(row[4]), row[5])

    def search(self, sql):
        """
        数据库查询
        :param sql: 数据库sql查询语句
        :return: 返回查询的条目
        """
        self.cur.execute(sql)
        return self.cur.fetchall()


def get_date_string(datetime=None):  # 把datetime类型转化为字符串
    format_ = "%Y-%m-%d"
    if datetime == None:
        today = datetime.date.today()  # type=datetime.date
        date = today.strftime(format_) # 转化为字符串
    else:
        date = datetime.strftime(format_)
    return date


""" 封装函数 """


def get_table(table_index):
    with Mysql_db(**db_connect) as db:  # db是__enter__方法的返回值, 即这个数据库对象
        table_name = db.table_dict[table_index]
        sql = "select * from %s" % (table_name)
        return db.search(sql)


def write_to_db(data, table_index):
    with Mysql_db(**db_connect) as db:  # db是__enter__方法的返回值, 即这个数据库对象
        if type(table_index) == int:
            table_index = [table_index]
        for i, index in enumerate(table_index):
            db.commit_db(data[i], index)


def execute_sql(sql):
    """
    执行任何sql语句
    :param sql:  sql 语句
    :return:  返回执行的结果
    """
    with Mysql_db(**db_connect) as db:  # db是__enter__方法的返回值, 即这个数据库对象
        return db.search(sql)


def create_database():
    conn = pymysql.connect(**db_connect)
    cur = conn.cursor()
    with open('./book_rental.sql', encoding='utf-8', mode='r') as f:
        sql_list = f.read().split(';')[:-1]
        for x in sql_list:
            if '\n' in x:
                x = x.replace('\n', ' ')
            if '  ' in x:
                x = x.replace('  ', '')
            sql_item = x + ';'
            print(sql_item)
            cur.execute(sql_item)

    try:
        conn.commit()  # 提交
        print("commit successfully!")
    except:
        conn.rollback()  # 如果发生错误时回滚
        print("your commit fails")
    finally:
        cur.close()
        conn.close()


if __name__ == '__main__':
    """
    Python 对 with 的处理还很聪明。基本思想是with所求值的对象必须有一个 __enter__() 方法，一个 __exit__() 方法。

    紧跟 with 后面的语句被求值后，返回对象的 __enter__() 方法被调用，这个方法的返回值将被赋值给as后面的变量。

    当 with 后面的代码块全部被执行完之后，将调用前面返回对象的 __exit__() 方法。
    """
    example_bookitem = ["9787533936020", "月亮与刘便士", "39.80", "2"]
    example_customerinfo = ["万月", "monthly"]

    sql = "select version()"
    print("Database is %s" % execute_sql(sql))

    # with Mysql_db(**db_connect) as db:  # db是__enter__方法的返回值, 即这个数据库对象
    # cursor = db.cur
    # cursor.execute("SELECT VERSION()")
    # data = cursor.fetchone()
    # print("Database version is %s" % data)
    # print("today is %s" % get_today())
    # db.insert_customerinfo(example_customerinfo[0], example_customerinfo[1])
    # print(db.search("SELECT * from customer_info"))

    """ 插入 book_item表 """
    # db.insert_bookitem(ISBN="9787533936020", bname="月亮与刘便士", price="39.80", inventory="2")
    # cursor.execute("SELECT * from book_item")
    # data = cursor.fetchone()
    # print("ISBN={0}, bname={1}, price={2}, inventory={3}".format(data[0], data[1], data[2], data[3]))

    """ 插入book_item表 100条数据 """
    # for i in range(100):
    #     db.insert_bookitem(ISBN=str(int(example_bookitem[0]) + i + 1), bname=example_bookitem[1],
    #                        price=example_bookitem[2],inventory=example_bookitem[3])
    # cursor.execute("SELECT * from book_item")
    # data = cursor.fetchone()
    # print("ISBN={0}, bname={1}, price={2}, inventory={3}".format(data[0], data[1], data[2], data[3]))

    """ 插入 customer_info表 """
    # db.insert_customerinfo("万力", "yearly")
    # cursor.execute("SELECT * from customer_info")
    # data = cursor.fetchone()
    # print("uid={0}, uname={1}, class={2}".format(data[0], data[1], data[2]))

    """ 插入 order_info 表 """
    # db.insert_orderinfo_start("1", get_today(), "月亮与六便士")
    # cursor.execute("SELECT * from order_info")
    # data = cursor.fetchall()
    # for i in data:
    #     print("uid={0}, start_time={1}, ISBN={2}".format(i[0], i[1], i[3]))

    """ 删除表customer_info中uid>1的记录 """
    # sql = "delete from customer_info where uid>1"
    # cursor.execute(sql)
    # cursor.execute("SELECT * from customer_info")
    # data = cursor.fetchall()
    # for i in data:
    #     print("uid={0}, uname={1}, class={2}".format(i[0], i[1], i[2]))

# " 读Excel表 "
# filename = "./书单示例.xlsx"
# data = read_excel(filename, 0)
# print(data)
#
# " 写Excel表 "
# writefile = "./写入实例.xlsx"
# sheet_name = " 书单信息表 "
# write_excel(writefile, sheet_name, data)
