import wx_sdk
import openpyxl
import json
import time
"读取ISBN号"

def read_excel(filename, table_index=0):
    data = []
    row_data = []
    workbook = openpyxl.load_workbook(filename)
    sheet_names = workbook.sheetnames
    worksheet = workbook.worksheets[table_index]  # 通过索引
    # 4. 获取工作表的属性
    sheet_name = worksheet.title
    rows = worksheet.max_row
    columns = worksheet.max_column
    # 5. 获取数据(sheet.rows 行生成器;sheet.columns列生成器)
    for row in worksheet.rows:
        for cell in row:
            row_data.append(cell.value)
        data.append(row_data)
        row_data = []

    return data


def write_excel(filename, data):
    workbook = openpyxl.Workbook()       # 创建工作簿
    worksheet = workbook.active          # 获取工作表
    row = len(data)
    column = len(data[0])
    for i in range(row):
        for j in range(column):
            worksheet.cell(i+1, j+1).value = data[i][j] # 写数据

    workbook.save(filename)              # 保存到文件
    return True

def delete_the_same(isbn_data):
    """
    删除有重复的isbn数据, 同时统计每本书的库存
    :param isbn_data: [[isbn1], [isbn2], ...]
    :return: [[isbn1, 2], [isbn2, 3], ...]
    """
    last_isbn = isbn_data[0][0]         # 第一个ISBN号
    num = 0                             # 该ISBN号有几个
    new_data = []
    for row in isbn_data:
        cur_isbn = row[0]
        if cur_isbn == last_isbn:
            num += 1
        else:
            new_data.append([last_isbn, num])
            last_isbn = cur_isbn        # 当前isbn成为下一个isbn的上一个isbn
            num = 1                     # 把新的isbn记录一次
    return new_data


def get_additional_data(*data):
    " data: [isbn, inventory]"
    isbn = data[0]
    inventory = data[1]
    url = 'https://way.jd.com/showapi/isbn'
    params = {
        'isbn' : isbn,
        'appkey' : '8b080ec6aae58e0eecf2f42572303851'
    }
    response = wx_sdk.wx_post_req( url, params )
    try:
        text = json.loads(response.text)
        info = text["result"]["showapi_res_body"]["data"]
        return [info['isbn'], info['title'], info['price'], inventory]
    except:
        print(response)
        return [isbn, "can't find the title", "no price", inventory]


if __name__ == "__main__":
    # isbn_filename = "./星月书店.xlsx"
    # isbn_data_with_same = read_excel(isbn_filename)
    # isbn_data_no_same = delete_the_same(isbn_data_with_same)

    isbn_data_no_same = read_excel("./星月书店书单完整信息.xlsx")
    # complete_data = []
    for row in isbn_data_no_same:
        # time.sleep(2)
        if row[2] == "no price":
            try:
                new_row = get_additional_data(row[0], row[3])
                for i in range(len(row)):
                    row[i] = new_row[i]
            except:
                pass
        # complete_data.append(get_additional_data(*row))
    write_excel("./星月书店书单完整信息.xlsx", isbn_data_no_same)
